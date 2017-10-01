# -*- coding: utf-8 -*-
"""
Created on Tue Jan 24 11:37:17 2017

@author: ckirst
"""


from openpyxl import load_workbook

wb = load_workbook(filename = '/home/ckirst/Science/Projects/CElegansBehaviour/Experiment/DwellingRoaming/strains vs N2 speed 1 bin per stage.xlsx')

wb = load_workbook(filename = '/home/ckirst/Science/Projects/CElegansBehaviour/Experiment/DwellingRoaming/strains vs N2 CV 1 bin per stage.xlsx')


stages = ['L1', 'L2', 'L3', 'L4', 'Adult'];
nstages = len(stages);
strains  = ['N2', 'tph-1', 'cat-2', 'tdc-1', 'npr-1', 'daf-7'];
nstrains = len(strains);
rows = ['A', 'B', 'C', 'D', 'E', 'F'];

data = {};
for ns, strain in enumerate(strains):
  strain = strains[ns];
  data[strain] = {};

  for nst, stage in enumerate(stages):
    ws = wb[stage];
    data[strain][stage] = [];
    for i in range(2, 128):
      v = ws[rows[ns] + str(i+1)].value;
      if v is not None:
        data[strain][stage].append(v);

data_all = {};
for strain in strains:
  n = len(data[strain]['L1']);
  data_all[strain] = np.zeros((nstages, n));
  for nst, stage in enumerate(stages):
    data_all[strain][nst,:] = data[strain][stage];


### plot some histograms
plt.figure(19); plt.clf();
for ns,strain in enumerate(strains):
  plt.subplot(3,2,ns+1);
  plt.hist(data_all[strain].T, bins  = 20)
  plt.title(strain)
  
  
### plot some histograms
plt.figure(19); plt.clf();
for ns,strain in enumerate(strains):
  plt.subplot(3,2,ns+1);
  plt.hist(data_all[strain].T, bins  = 20, range = (0, 1))
  plt.title(strain)
  
  
dbins = 7;
tbins = 5;
max_speed = 0.8;

dist_t = {};
for strain in strains:
  dist_t[strain] = np.zeros((dbins, tbins));
  for t in range(tbins):
    dist_t[strain][:,t] = np.histogram(data_all[strain][t,:], range = (0.2, max_speed), bins = dbins)[0];
    dist_t[strain][:,t] /= np.sum(dist_t[strain][:,t]);
    
    
### plot some histograms
plt.figure(99); plt.clf();
for ns,strain in enumerate(strains):
  plt.subplot(3,2,ns+1);
  plt.imshow(dist_t[strain].T, aspect = 'auto')
  plt.title(strain)
plt.tight_layout();

